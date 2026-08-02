from __future__ import annotations

import base64
import csv
import json
import mimetypes
import os
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .models import ExtractionBatch, RawEntry


HEADER_ALIASES = {
    "date": {"date", "tanggal"},
    "airline": {"airline", "operator"},
    "flight_number": {"flight number", "flight no", "flt no", "flight"},
    "route": {"route", "routing", "from-to", "sector"},
    "out_time": {"out", "out time"},
    "in_time": {"in", "in time"},
    "total_time": {"total time", "block time", "flight time", "total"},
    "registration": {"registration", "aircraft registration", "reg"},
    "aircraft_type": {"aircraft type", "type"},
    "pic_name": {"pic", "pic name", "captain", "captain name"},
    "pic_employee_id": {"pic id", "employee id", "captain id"},
    "p1_us": {"p1 u/s", "p1us"},
    "simulator_time": {"simulator time", "sim time"},
    "approach": {"approach", "approach type"},
}


class StructuredFileExtractor:
    def extract(self, path: Path) -> ExtractionBatch:
        if path.suffix.lower() == ".json":
            return ExtractionBatch.model_validate_json(path.read_text(encoding="utf-8"))
        if path.suffix.lower() == ".csv":
            with path.open(newline="", encoding="utf-8-sig") as handle:
                rows = list(csv.DictReader(handle))
            return ExtractionBatch(entries=[self._from_row(row, path.name, idx + 2) for idx, row in enumerate(rows)])
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            return self._from_xlsx(path)
        raise ValueError(f"Unsupported structured source: {path}")

    def _from_xlsx(self, path: Path) -> ExtractionBatch:
        wb = load_workbook(path, data_only=True, read_only=False)
        entries: list[RawEntry] = []
        for ws in wb.worksheets:
            header_row, mapping = self._detect_header(ws)
            if not mapping:
                continue
            for row_idx in range(header_row + 1, ws.max_row + 1):
                row = {field: ws.cell(row_idx, col_idx).value for field, col_idx in mapping.items()}
                if not any(value not in (None, "") for value in row.values()):
                    continue
                entries.append(self._from_row(row, path.name, f"{ws.title}!{row_idx}"))
        return ExtractionBatch(entries=entries)

    def _detect_header(self, ws) -> tuple[int, dict[str, int]]:
        for row_idx in range(1, min(ws.max_row, 30) + 1):
            mapping: dict[str, int] = {}
            for col_idx in range(1, ws.max_column + 1):
                value = str(ws.cell(row_idx, col_idx).value or "").strip().casefold()
                for field, aliases in HEADER_ALIASES.items():
                    if value in aliases:
                        mapping[field] = col_idx
            if {"date", "route"}.issubset(mapping):
                return row_idx, mapping
        return 0, {}

    def _from_row(self, row: dict, source_file: str, row_number: object) -> RawEntry:
        normalized: dict[str, object] = {}
        for key, value in row.items():
            key_cf = str(key).strip().casefold()
            target = next((field for field, aliases in HEADER_ALIASES.items() if key_cf in aliases or key_cf == field), None)
            if target:
                normalized[target] = value
        return RawEntry(
            source_file=source_file,
            source_row=str(row_number),
            date=_as_date(normalized.get("date")),
            airline=_as_text(normalized.get("airline")),
            flight_number=_as_text(normalized.get("flight_number")),
            route_sequence=_as_text(normalized.get("route")) or "",
            out_time=_as_clock(normalized.get("out_time")),
            in_time=_as_clock(normalized.get("in_time")),
            total_time=_as_duration(normalized.get("total_time")),
            registration=_as_text(normalized.get("registration")),
            aircraft_type=_as_text(normalized.get("aircraft_type")),
            pic_name=_as_text(normalized.get("pic_name")),
            pic_employee_id=_as_text(normalized.get("pic_employee_id")),
            p1_us=_as_duration(normalized.get("p1_us")),
            simulator_time=_as_duration(normalized.get("simulator_time")),
            approach=_as_text(normalized.get("approach")),
        )


class OpenAIVisionExtractor:
    def __init__(self, model: str | None = None):
        if not os.getenv("OPENAI_API_KEY"):
            raise RuntimeError("OPENAI_API_KEY is required for image/PDF extraction")
        from openai import OpenAI

        self.client = OpenAI()
        self.model = model or os.getenv("OPENAI_MODEL", "gpt-5")

    def extract(self, path: Path) -> ExtractionBatch:
        content = [
            {
                "type": "input_text",
                "text": (
                    "Extract Lion Air pilot-logbook entries from this source. Preserve readable values exactly. "
                    "Do not guess unreadable fields. Return route_sequence as ordered airport codes. "
                    "Use HH:MM for clock and duration fields. Mark unreadable field names in unreadable_fields. "
                    "Do not process another airline."
                ),
            },
            self._file_content(path),
        ]
        schema = ExtractionBatch.model_json_schema()
        response = self.client.responses.create(
            model=self.model,
            input=[{"role": "user", "content": content}],
            text={
                "format": {
                    "type": "json_schema",
                    "name": "lion_air_logbook_extraction",
                    "strict": False,
                    "schema": schema,
                }
            },
        )
        return ExtractionBatch.model_validate_json(response.output_text)

    def _file_content(self, path: Path) -> dict:
        mime = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        encoded = base64.b64encode(path.read_bytes()).decode("ascii")
        if mime == "application/pdf":
            return {
                "type": "input_file",
                "filename": path.name,
                "file_data": f"data:{mime};base64,{encoded}",
            }
        return {
            "type": "input_image",
            "detail": "high",
            "image_url": f"data:{mime};base64,{encoded}",
        }


def extractor_for(path: Path):
    if path.suffix.lower() in {".json", ".csv", ".xlsx", ".xlsm"}:
        return StructuredFileExtractor()
    if path.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp", ".pdf"}:
        return OpenAIVisionExtractor()
    raise ValueError(f"Unsupported source type: {path.suffix}")


def _as_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text or None


def _as_date(value: object):
    if value is None:
        return None
    if hasattr(value, "date"):
        return value.date()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _as_clock(value: object) -> str | None:
    if value is None:
        return None
    if hasattr(value, "hour") and hasattr(value, "minute"):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, (int, float)) and 0 <= float(value) < 1:
        minutes = round(float(value) * 24 * 60)
        return f"{minutes // 60:02d}:{minutes % 60:02d}"
    return _as_text(value)


def _as_duration(value: object) -> str | None:
    if value is None:
        return None
    if hasattr(value, "total_seconds"):
        minutes = round(value.total_seconds() / 60)
        return f"{minutes // 60:02d}:{minutes % 60:02d}"
    if isinstance(value, (int, float)) and 0 <= float(value) < 10:
        minutes = round(float(value) * 24 * 60)
        return f"{minutes // 60:02d}:{minutes % 60:02d}"
    return _as_text(value)
