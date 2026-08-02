from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter

from .models import AirportMapping, FlightSector, ProvenanceRecord, SourceInventoryItem
from .time_utils import excel_duration, parse_duration_minutes
from .validation import ValidationResult


MAIN_HEADERS = [
    "No.",
    "Date",
    "Aircraft Type",
    "Registration",
    "Flight No.",
    "From",
    "To",
    "OUT",
    "IN",
    "Total Time",
    "PIC Name",
    "PIC ID",
    "SIC Name",
    "P1 U/S",
    "Copilot Time",
    "IFR",
    "ACTUAL IFR",
    "Simulator Time",
    "Approach",
    "Remark",
]


class WorkbookWriter:
    def __init__(self, owner_name: str, rows_per_page: int = 24):
        self.owner_name = owner_name
        self.rows_per_page = rows_per_page
        self.thin = Side(style="thin", color="000000")
        self.medium = Side(style="medium", color="000000")

    def write(
        self,
        destination: Path,
        sectors: list[FlightSector],
        provenance: list[ProvenanceRecord],
        validation: ValidationResult,
        mappings: list[AirportMapping],
        inventory: list[SourceInventoryItem],
    ) -> Path:
        destination.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "INTERNATIONAL CORADINE"
        self._write_main(ws, sectors)
        self._write_provenance(wb.create_sheet("DATA PROVENANCE"), provenance)
        self._write_validation(wb.create_sheet("VALIDATION REPORT"), validation)
        self._write_mapping(wb.create_sheet("AIRPORT CODE MAPPING"), mappings)
        self._write_inventory(wb.create_sheet("SOURCE INVENTORY"), inventory)
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
        wb.save(destination)
        return destination

    def _write_main(self, ws, sectors: list[FlightSector]) -> None:
        ws.merge_cells("A1:T1")
        ws["A1"] = "INTERNATIONAL CORADINE LOGBOOK — REV 3"
        ws["A1"].font = Font(name="Arial", size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        ws.merge_cells("A2:T2")
        ws["A2"] = f"LOGBOOK OWNER: {self.owner_name} | LION AIR PILOT USE ONLY"
        ws["A2"].font = Font(name="Arial", size=10, bold=True)
        ws["A2"].alignment = Alignment(horizontal="center")

        groups = [
            ("A3:B3", "ENTRY"),
            ("C3:D3", "AIRCRAFT"),
            ("E3:G3", "FLIGHT / ROUTE"),
            ("H3:J3", "TIMES"),
            ("K3:M3", "CREW"),
            ("N3:O3", "CAPACITY"),
            ("P3:Q3", "INSTRUMENT"),
            ("R3:T3", "OTHER"),
        ]
        for cell_range, label in groups:
            ws.merge_cells(cell_range)
            cell = ws[cell_range.split(":")[0]]
            cell.value = label
            cell.font = Font(name="Arial", size=8, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="D9EAF7")

        for col, header in enumerate(MAIN_HEADERS, start=1):
            cell = ws.cell(4, col, header)
            cell.font = Font(name="Arial", size=8, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor="EAF2F8")
            cell.border = Border(top=self.medium, bottom=self.medium, left=self.thin, right=self.thin)
        ws.row_dimensions[4].height = 32

        start_row = 5
        for offset, sector in enumerate(sectors):
            row = start_row + offset
            values = [
                sector.entry_number,
                sector.date,
                sector.aircraft_type,
                sector.registration,
                sector.flight_number,
                sector.departure,
                sector.arrival,
                sector.out_time,
                sector.in_time,
                excel_duration(parse_duration_minutes(sector.total_time)),
                sector.pic_name,
                sector.pic_employee_id,
                sector.sic_name,
                excel_duration(parse_duration_minutes(sector.p1_us)),
                excel_duration(parse_duration_minutes(sector.copilot_time)),
                f"=J{row}" if sector.total_time != "—" else "—",
                f"=J{row}" if sector.total_time != "—" else "—",
                excel_duration(parse_duration_minutes(sector.simulator_time)),
                sector.approach,
                "",
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row, col, value)
                cell.font = Font(name="Arial", size=7)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)
                if col in {10, 14, 15, 16, 17, 18} and isinstance(value, (float, int)):
                    cell.number_format = "[h]:mm"
                if col in {16, 17} and isinstance(value, str) and value.startswith("="):
                    cell.number_format = "[h]:mm"
                    cell.protection = Protection(locked=True)
                else:
                    cell.protection = Protection(locked=False)
            ws.row_dimensions[row].height = 24
            if offset and offset % self.rows_per_page == 0:
                ws.row_breaks.append(Break(id=row - 1))

        total_row = start_row + len(sectors)
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=9)
        ws.cell(total_row, 1, "PAGE / WORKBOOK TOTAL")
        ws.cell(total_row, 10, f"=SUM(J{start_row}:J{total_row-1})")
        ws.cell(total_row, 14, f"=SUM(N{start_row}:N{total_row-1})")
        ws.cell(total_row, 15, f"=SUM(O{start_row}:O{total_row-1})")
        ws.cell(total_row, 16, f"=SUM(P{start_row}:P{total_row-1})")
        ws.cell(total_row, 17, f"=SUM(Q{start_row}:Q{total_row-1})")
        ws.cell(total_row, 18, f"=SUM(R{start_row}:R{total_row-1})")
        for col in range(1, 21):
            cell = ws.cell(total_row, col)
            cell.font = Font(name="Arial", size=8, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=self.medium, bottom=self.medium, left=self.thin, right=self.thin)
            if col in {10, 14, 15, 16, 17, 18}:
                cell.number_format = "[h]:mm"

        widths = [5, 11, 14, 13, 10, 8, 8, 7, 7, 10, 22, 11, 22, 9, 11, 9, 11, 12, 15, 12]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:T{max(4, total_row-1)}"
        ws.print_title_rows = "1:4"
        ws.print_area = f"A1:T{total_row}"
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A3
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.oddFooter.center.text = "Personal analytical reconstruction — not an officially certified logbook"
        ws.oddFooter.right.text = "Page &P of &N"
        ws.page_margins.left = 0.2
        ws.page_margins.right = 0.2
        ws.page_margins.top = 0.3
        ws.page_margins.bottom = 0.3
        ws.protection.sheet = True
        ws.protection.selectLockedCells = False
        ws.protection.selectUnlockedCells = True

        ws.conditional_formatting.add(
            f"A5:T{max(5, total_row-1)}",
            FormulaRule(formula=["$F5=$G5"], fill=PatternFill("solid", fgColor="FFF2CC")),
        )

    def _write_provenance(self, ws, records: list[ProvenanceRecord]) -> None:
        headers = [
            "Entry number",
            "Date",
            "Route",
            "Field name",
            "Final value",
            "Classification",
            "Source used",
            "Reasoning or calculation",
            "Website or document reference",
            "Access date",
            "Confidence level",
            "Any unresolved issue",
        ]
        rows = [
            [
                r.entry_number,
                r.date,
                r.route,
                r.field_name,
                r.final_value,
                r.classification.value,
                r.source_used,
                r.reasoning_or_calculation,
                r.website_or_document_reference,
                r.access_date,
                r.confidence_level,
                r.unresolved_issue,
            ]
            for r in records
        ]
        self._write_table(ws, headers, rows, widths=[12, 12, 18, 20, 18, 14, 38, 55, 40, 12, 14, 45])

    def _write_validation(self, ws, validation: ValidationResult) -> None:
        ws["A1"] = "VALIDATION SUMMARY"
        ws["A1"].font = Font(size=14, bold=True)
        row = 3
        for key, value in validation.summary.items():
            ws.cell(row, 1, key)
            ws.cell(row, 2, value)
            row += 1
        row += 1
        ws.cell(row, 1, "ISSUES")
        ws.cell(row, 1).font = Font(bold=True)
        row += 1
        headers = ["Severity", "Entry number", "Code", "Message"]
        for col, header in enumerate(headers, 1):
            ws.cell(row, col, header)
        for issue in validation.issues:
            row += 1
            for col, value in enumerate([issue.severity, issue.entry_number, issue.code, issue.message], 1):
                ws.cell(row, col, value)
        self._style_used_range(ws)
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 26
        ws.column_dimensions["D"].width = 85
        ws.freeze_panes = "A3"

    def _write_mapping(self, ws, mappings: list[AirportMapping]) -> None:
        headers = [
            "Original airport code",
            "Original code type",
            "Final ICAO code",
            "Airport name",
            "Country",
            "Historical code",
            "Current code",
            "Source",
            "Access date",
            "Notes",
        ]
        rows = [
            [
                m.original_airport_code,
                m.original_code_type,
                m.final_icao_code,
                m.airport_name,
                m.country,
                m.historical_code,
                m.current_code,
                m.source,
                m.access_date,
                m.notes,
            ]
            for m in mappings
        ]
        self._write_table(ws, headers, rows, widths=[22, 20, 18, 42, 18, 18, 18, 55, 14, 45])

    def _write_inventory(self, ws, inventory: list[SourceInventoryItem]) -> None:
        headers = [
            "Source number",
            "Filename",
            "File type",
            "Number of pages or sheets",
            "Apparent date range",
            "Source priority",
            "Description",
            "Readability status",
            "Whether used",
            "SHA-256",
            "Notes",
        ]
        rows = [
            [
                item.source_number,
                item.filename,
                item.file_type,
                item.pages_or_sheets,
                item.apparent_date_range,
                item.source_priority,
                item.description,
                item.readability_status,
                "Yes" if item.used else "No",
                item.sha256,
                item.notes,
            ]
            for item in inventory
        ]
        self._write_table(ws, headers, rows, widths=[14, 35, 12, 24, 28, 16, 45, 24, 14, 70, 45])

    def _write_table(self, ws, headers: list[str], rows: list[list], widths: list[int]) -> None:
        for col, header in enumerate(headers, start=1):
            ws.cell(1, col, header)
        for row_idx, values in enumerate(rows, start=2):
            for col, value in enumerate(values, start=1):
                ws.cell(row_idx, col, value)
        self._style_used_range(ws)
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows)+1)}"
        ws.sheet_view.showGridLines = False

    def _style_used_range(self, ws) -> None:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name="Arial", size=9, bold=cell.row == 1)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)
                if cell.row == 1:
                    cell.fill = PatternFill("solid", fgColor="D9EAF7")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
