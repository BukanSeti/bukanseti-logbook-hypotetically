from __future__ import annotations

import shutil
import subprocess
import tempfile
from pathlib import Path

from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, TableStyle

from .models import FlightSector


def export_pdf_from_workbook(
    workbook_path: Path,
    pdf_path: Path,
    sectors: list[FlightSector],
    owner_name: str,
    allow_fallback: bool = False,
) -> Path:
    office = shutil.which("libreoffice") or shutil.which("soffice")
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    if office:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory(prefix="coradine-pdf-") as temporary_directory:
            temporary_directory_path = Path(temporary_directory)
            printable_workbook = temporary_directory_path / workbook_path.name
            workbook = load_workbook(workbook_path, data_only=False)
            for sheet_name in list(workbook.sheetnames):
                if sheet_name != "INTERNATIONAL CORADINE":
                    del workbook[sheet_name]
            workbook.save(printable_workbook)
            completed = subprocess.run(
                [office, "--headless", "--convert-to", "pdf", "--outdir", str(temporary_directory_path), str(printable_workbook)],
                check=False,
                capture_output=True,
                text=True,
                timeout=180,
            )
            generated = temporary_directory_path / f"{printable_workbook.stem}.pdf"
            if completed.returncode == 0 and generated.exists():
                shutil.copy2(generated, pdf_path)
                inspect_pdf(pdf_path)
                return pdf_path
            if not allow_fallback:
                raise RuntimeError(f"LibreOffice PDF export failed: {completed.stderr or completed.stdout}")
    elif not allow_fallback:
        raise RuntimeError(
            "LibreOffice/soffice is required for strict PDF-from-Excel export. "
            "Use the Docker image or pass --allow-pdf-fallback for a non-identical renderer."
        )

    _reportlab_fallback(pdf_path, sectors, owner_name)
    inspect_pdf(pdf_path)
    return pdf_path


def inspect_pdf(pdf_path: Path) -> None:
    import fitz

    with fitz.open(pdf_path) as document:
        if document.page_count < 1:
            raise RuntimeError("PDF contains no pages")
        for page_number, page in enumerate(document, start=1):
            if not page.get_text("text").strip():
                raise RuntimeError(f"PDF page {page_number} is unexpectedly blank")
            page_rect = page.rect
            for block in page.get_text("blocks"):
                x0, y0, x1, y1 = block[:4]
                if x0 < -1 or y0 < -1 or x1 > page_rect.width + 1 or y1 > page_rect.height + 1:
                    raise RuntimeError(f"Text block exceeds page bounds on page {page_number}")


def _reportlab_fallback(pdf_path: Path, sectors: list[FlightSector], owner_name: str) -> None:
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        str(pdf_path),
        pagesize=landscape(A3),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
        title="International Coradine Rev 3",
        author=owner_name,
    )
    headers = [
        "No", "Date", "Type", "Reg", "Flight", "From", "To", "OUT", "IN", "Total",
        "PIC", "PIC ID", "SIC", "P1 U/S", "Copilot", "IFR", "ACTUAL IFR", "SIM", "Approach", "Remark"
    ]
    data = [headers]
    for s in sectors:
        data.append([
            s.entry_number,
            s.date.isoformat() if s.date else "—",
            s.aircraft_type,
            s.registration,
            s.flight_number,
            s.departure,
            s.arrival,
            s.out_time,
            s.in_time,
            s.total_time,
            s.pic_name,
            s.pic_employee_id,
            s.sic_name,
            s.p1_us,
            s.copilot_time,
            s.ifr,
            s.actual_ifr,
            s.simulator_time,
            s.approach,
            "",
        ])
    table = LongTable(data, repeatRows=1, colWidths=[9*mm, 18*mm, 24*mm, 20*mm, 17*mm, 14*mm, 14*mm, 13*mm, 13*mm, 17*mm, 35*mm, 18*mm, 35*mm, 15*mm, 18*mm, 15*mm, 18*mm, 17*mm, 24*mm, 18*mm])
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.35, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 5.5),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 1.5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1.5),
    ]))
    story = [
        Paragraph("INTERNATIONAL CORADINE LOGBOOK — REV 3", styles["Title"]),
        Paragraph(f"Logbook owner: {owner_name} — Lion Air pilot use only", styles["Normal"]),
        Paragraph("Personal analytical reconstruction; not an officially certified pilot logbook.", styles["Italic"]),
        table,
    ]
    document.build(story)
