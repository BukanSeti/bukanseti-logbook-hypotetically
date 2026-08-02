from __future__ import annotations

import json
import os
import time
from dataclasses import dataclass
from io import BytesIO
from threading import Lock
from typing import Protocol
from urllib.parse import quote

from openpyxl import load_workbook


@dataclass(frozen=True)
class CrewLookupResult:
    employee_id: str
    full_name: str


@dataclass(frozen=True)
class AircraftLookupResult:
    registration: str
    aircraft_type: str


class ReferenceStore(Protocol):
    def lookup_crew(
        self,
        employee_id: str | None,
        full_name: str | None,
    ) -> CrewLookupResult | None: ...

    def lookup_aircraft(self, registration: str) -> AircraftLookupResult | None: ...


class GoogleDriveWorkbookExporter:
    def __init__(
        self,
        service_account_file: str | None = None,
        service_account_json: str | None = None,
    ):
        self.service_account_file = service_account_file
        self.service_account_json = service_account_json

    @classmethod
    def from_environment(cls) -> "GoogleDriveWorkbookExporter":
        return cls(
            service_account_file=os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE"),
            service_account_json=os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON"),
        )

    def export_xlsx(self, file_id: str) -> bytes:
        try:
            from google.auth.transport.requests import AuthorizedSession
            from google.oauth2 import service_account
        except ImportError as exc:
            raise RuntimeError(
                "Install the reference-api dependencies to use Google Drive"
            ) from exc

        scopes = ["https://www.googleapis.com/auth/drive.readonly"]
        if self.service_account_json:
            info = json.loads(self.service_account_json)
            credentials = service_account.Credentials.from_service_account_info(
                info,
                scopes=scopes,
            )
        elif self.service_account_file:
            credentials = service_account.Credentials.from_service_account_file(
                self.service_account_file,
                scopes=scopes,
            )
        else:
            raise RuntimeError("Google service-account credentials are not configured")

        session = AuthorizedSession(credentials)
        mime_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        url = (
            "https://www.googleapis.com/drive/v3/files/"
            f"{quote(file_id, safe='')}/export"
        )
        response = session.get(url, params={"mimeType": mime_type}, timeout=60)
        response.raise_for_status()
        return response.content


class GoogleSheetReferenceStore:
    def __init__(
        self,
        exporter: GoogleDriveWorkbookExporter,
        crew_file_id: str,
        aircraft_file_id: str,
        cache_ttl_seconds: int = 300,
    ):
        self.exporter = exporter
        self.crew_file_id = crew_file_id
        self.aircraft_file_id = aircraft_file_id
        self.cache_ttl_seconds = cache_ttl_seconds
        self._loaded_at = 0.0
        self._crew_by_id: dict[str, CrewLookupResult] = {}
        self._crew_by_name: dict[str, CrewLookupResult] = {}
        self._aircraft: dict[str, AircraftLookupResult] = {}
        self._lock = Lock()

    @classmethod
    def from_environment(cls) -> "GoogleSheetReferenceStore":
        crew_file_id = os.getenv("LION_AIR_CREW_SHEET_ID", "").strip()
        aircraft_file_id = os.getenv("LION_AIR_AIRCRAFT_SHEET_ID", "").strip()
        if not crew_file_id or not aircraft_file_id:
            raise RuntimeError("Both private Google Sheet file IDs must be configured")
        return cls(
            exporter=GoogleDriveWorkbookExporter.from_environment(),
            crew_file_id=crew_file_id,
            aircraft_file_id=aircraft_file_id,
            cache_ttl_seconds=int(
                os.getenv("CORADINE_REFERENCE_CACHE_TTL_SECONDS", "300")
            ),
        )

    def _ensure_loaded(self) -> None:
        now = time.monotonic()
        if self._loaded_at and now - self._loaded_at < self.cache_ttl_seconds:
            return
        with self._lock:
            now = time.monotonic()
            if self._loaded_at and now - self._loaded_at < self.cache_ttl_seconds:
                return
            crew_bytes = self.exporter.export_xlsx(self.crew_file_id)
            aircraft_bytes = self.exporter.export_xlsx(self.aircraft_file_id)
            crew_by_id, crew_by_name = _load_crew(crew_bytes)
            aircraft = _load_aircraft(aircraft_bytes)
            self._crew_by_id = crew_by_id
            self._crew_by_name = crew_by_name
            self._aircraft = aircraft
            self._loaded_at = now

    def lookup_crew(
        self,
        employee_id: str | None,
        full_name: str | None,
    ) -> CrewLookupResult | None:
        self._ensure_loaded()
        if employee_id:
            return self._crew_by_id.get(_normalize_id(employee_id))
        if full_name:
            return self._crew_by_name.get(_clean_name(full_name).casefold())
        return None

    def lookup_aircraft(self, registration: str) -> AircraftLookupResult | None:
        self._ensure_loaded()
        return self._aircraft.get(_normalize_registration(registration))


def _load_crew(
    content: bytes,
) -> tuple[dict[str, CrewLookupResult], dict[str, CrewLookupResult]]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    worksheet = workbook.active
    headers, start_row = _headers(worksheet, "Airline")
    by_id: dict[str, CrewLookupResult] = {}
    by_name: dict[str, CrewLookupResult] = {}
    for row in worksheet.iter_rows(min_row=start_row, values_only=True):
        record = dict(zip(headers, row, strict=False))
        if str(record.get("Airline") or "").strip().upper() != "LION AIR":
            continue
        employee_id = _normalize_id(record.get("Employee ID"))
        full_name = _clean_name(str(record.get("Crew Name") or ""))
        if not employee_id or not full_name:
            continue
        item = CrewLookupResult(employee_id=employee_id, full_name=full_name)
        by_id[employee_id] = item
        by_name[full_name.casefold()] = item
    return by_id, by_name


def _load_aircraft(content: bytes) -> dict[str, AircraftLookupResult]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    worksheet = workbook.active
    headers, start_row = _headers(worksheet, "Aircraft Registration")
    output: dict[str, AircraftLookupResult] = {}
    for row in worksheet.iter_rows(min_row=start_row, values_only=True):
        record = dict(zip(headers, row, strict=False))
        registration = _normalize_registration(
            str(record.get("Aircraft Registration") or "")
        )
        aircraft_type = _standardize_type(
            str(
                record.get("Type of Aircraft")
                or record.get("Aircraft Type (ICAO)")
                or ""
            )
        )
        if registration and aircraft_type:
            output[registration] = AircraftLookupResult(
                registration=registration,
                aircraft_type=aircraft_type,
            )
    return output


def _headers(worksheet, required_header: str) -> tuple[list[str], int]:
    for row_number in range(1, min(worksheet.max_row, 20) + 1):
        row = next(
            worksheet.iter_rows(
                min_row=row_number,
                max_row=row_number,
                values_only=True,
            )
        )
        values = [str(value or "").strip() for value in row]
        if required_header in values:
            return values, row_number + 1
    raise RuntimeError(f"Required header {required_header!r} was not found")


def _normalize_id(value: object) -> str:
    text = str(value or "").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _clean_name(value: str) -> str:
    text = " ".join(value.strip().split())
    upper = text.upper()
    for prefix in ("CAPTAIN ", "CAPT ", "CPT "):
        if upper.startswith(prefix):
            return text[len(prefix) :].strip()
    return text


def _normalize_registration(value: str) -> str:
    text = value.strip().upper().replace(" ", "")
    if text and not text.startswith("PK-") and len(text) == 3:
        text = f"PK-{text}"
    return text


def _standardize_type(value: str) -> str:
    normalized = value.strip().upper().replace("_", "-")
    if normalized in {"B7379", "B739", "737-900ER", "BOEING 737-900ER"}:
        return "B737-900ER"
    if normalized in {
        "B7378",
        "B738",
        "737-800",
        "737-800NG",
        "BOEING 737-800",
        "BOEING 737-800NG",
    }:
        return "B737-800NG"
    return value.strip()
