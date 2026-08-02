from __future__ import annotations

import os
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import requests
from openpyxl import load_workbook

from .lion_air_guard import assert_lion_air_operator


@dataclass(frozen=True)
class CrewRecord:
    airline: str
    name: str
    employee_id: str | None
    atpl_number: str | None
    rank_or_role: str | None
    employment_status: str | None
    validation_status: str | None


@dataclass(frozen=True)
class AircraftRecord:
    registration: str
    icao_type: str | None
    type_of_aircraft: str | None
    variant: str | None
    operator: str | None
    validation_status: str | None


def sheet_export_url(url: str) -> str:
    if "/d/" not in url:
        raise ValueError("Invalid Google Sheets URL")
    spreadsheet_id = url.split("/d/", 1)[1].split("/", 1)[0]
    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"


def download_sheet_xlsx(url: str, destination: Path) -> Path:
    destination.parent.mkdir(parents=True, exist_ok=True)
    response = requests.get(sheet_export_url(url), timeout=60)
    response.raise_for_status()
    destination.write_bytes(response.content)
    return destination


def _find_header_row(ws, required: str) -> int:
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        values = [str(ws.cell(row_idx, col).value or "").strip() for col in range(1, ws.max_column + 1)]
        if required in values:
            return row_idx
    raise ValueError(f"Header {required!r} not found in {ws.title}")


def _rows_as_dicts(path: Path, required_header: str) -> Iterable[dict[str, object]]:
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb.active
    header_row = _find_header_row(ws, required_header)
    headers = [str(ws.cell(header_row, col).value or "").strip() for col in range(1, ws.max_column + 1)]
    for row_idx in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if not any(value not in (None, "") for value in values):
            continue
        yield dict(zip(headers, values, strict=False))


class CrewBank:
    def __init__(self, records: list[CrewRecord]):
        self.records = [record for record in records if record.airline.strip().upper() == "LION AIR"]
        self.by_name = {record.name.casefold(): record for record in self.records}
        self.by_id = {record.employee_id: record for record in self.records if record.employee_id}

    @classmethod
    def from_xlsx(cls, path: Path) -> "CrewBank":
        records = []
        for row in _rows_as_dicts(path, "Airline"):
            airline = str(row.get("Airline") or "").strip()
            if airline.upper() != "LION AIR":
                continue
            records.append(
                CrewRecord(
                    airline=airline,
                    name=str(row.get("Crew Name") or "").strip(),
                    employee_id=_text(row.get("Employee ID")),
                    atpl_number=_text(row.get("ATPL Number")),
                    rank_or_role=_text(row.get("Rank or Role")),
                    employment_status=_text(row.get("Employment Status")),
                    validation_status=_text(row.get("Validation Status")),
                )
            )
        return cls(records)

    def lookup(self, name: str | None = None, employee_id: str | None = None) -> CrewRecord | None:
        if employee_id and employee_id in self.by_id:
            return self.by_id[employee_id]
        if name:
            return self.by_name.get(_clean_rank(name).casefold())
        return None


class AircraftBank:
    def __init__(self, records: list[AircraftRecord]):
        self.records = records
        self.by_registration = {record.registration.upper(): record for record in records}

    @classmethod
    def from_xlsx(cls, path: Path) -> "AircraftBank":
        records = []
        for row in _rows_as_dicts(path, "Aircraft Registration"):
            operator = _text(row.get("Operator"))
            if operator:
                assert_lion_air_operator(operator)
            records.append(
                AircraftRecord(
                    registration=str(row.get("Aircraft Registration") or "").strip().upper(),
                    icao_type=_text(row.get("Aircraft Type (ICAO)")),
                    type_of_aircraft=_text(row.get("Type of Aircraft")),
                    variant=_text(row.get("Aircraft Variant")),
                    operator=operator,
                    validation_status=_text(row.get("Validation Status")),
                )
            )
        return cls(records)

    def lookup(self, registration: str | None) -> AircraftRecord | None:
        if not registration:
            return None
        normalized = registration.strip().upper()
        if normalized and not normalized.startswith("PK-") and len(normalized) == 3:
            normalized = f"PK-{normalized}"
        return self.by_registration.get(normalized)


def resolve_reference_path(configured: str, environment_variable: str) -> Path:
    return Path(os.getenv(environment_variable, configured)).expanduser()


def _text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text or None


def _clean_rank(name: str) -> str:
    cleaned = name.strip()
    for prefix in ("CAPTAIN ", "CAPT ", "CPT "):
        if cleaned.upper().startswith(prefix):
            return cleaned[len(prefix) :].strip()
    return cleaned
