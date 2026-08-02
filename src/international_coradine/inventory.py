from __future__ import annotations

import hashlib
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .models import SourceInventoryItem


def build_source_inventory(paths: list[Path]) -> list[SourceInventoryItem]:
    items = []
    for index, path in enumerate(paths, start=1):
        suffix = path.suffix.lower().lstrip(".") or "unknown"
        pages_or_sheets = "—"
        description = "Uploaded source"
        if suffix in {"xlsx", "xlsm"}:
            wb = load_workbook(path, read_only=True, data_only=True)
            pages_or_sheets = str(len(wb.sheetnames))
            description = f"Spreadsheet: {', '.join(wb.sheetnames)}"
        elif suffix == "pdf":
            try:
                import fitz

                with fitz.open(path) as doc:
                    pages_or_sheets = str(doc.page_count)
            except Exception:
                pages_or_sheets = "Unknown"
        elif suffix in {"png", "jpg", "jpeg", "webp"}:
            pages_or_sheets = "1"
            description = "Logbook photograph or scan"
        items.append(
            SourceInventoryItem(
                source_number=index,
                filename=path.name,
                file_type=suffix.upper(),
                pages_or_sheets=pages_or_sheets,
                apparent_date_range="To be determined during extraction",
                source_priority=_priority(suffix),
                description=description,
                readability_status="Pending extraction",
                used=False,
                sha256=hashlib.sha256(path.read_bytes()).hexdigest(),
                notes=f"Inventoried {datetime.now().isoformat(timespec='seconds')}",
            )
        )
    return items


def _priority(suffix: str) -> int:
    if suffix in {"png", "jpg", "jpeg", "webp"}:
        return 1
    if suffix == "pdf":
        return 2
    if suffix in {"xlsx", "xlsm", "csv", "json"}:
        return 3
    return 9
